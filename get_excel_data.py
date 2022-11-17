from openpyxl import load_workbook

excel_path = r'/Users/elenalysenko/Desktop/QA/PyTest/demo_swag/data/data.xlsx'
wb = load_workbook(excel_path)

ws = wb.worksheets[0]

login_form_parameters = []
for i in range(2, ws.max_row+1):
    username = ws.cell(column=1, row=i).value
    password = ws.cell(column=2, row=i).value
    checkpoint = ws.cell(column=3, row=i).value
    tpl = username, password, checkpoint
    login_form_parameters.append(tpl)
